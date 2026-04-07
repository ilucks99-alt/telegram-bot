import math
import re
import shutil
import tempfile
import pandas as pd
import logging
import config
import os
import json

from telegram_service import send_message
from typing import Any, Dict, Optional, Tuple
from datetime import datetime
from zoneinfo import ZoneInfo

# =========================================================
# 유틸
# =========================================================
def normalize_text(s: Any) -> str:
    if s is None or pd.isna(s):
        return ""
    s = str(s).strip().lower()
    s = re.sub(r"[\s\-\_\(\)\[\]\{\}\.,/&|:;]+", "", s)
    return s


def safe_num(x: Any) -> Optional[float]:
    if pd.isna(x):
        return None
    try:
        val = float(x)
        if math.isnan(val):
            return None
        return val
    except Exception:
        return None


def format_amount_uk(value: Optional[float]) -> str:
    if value is None:
        return "N/A"
    return f"{value:,.0f}억"


def format_pct(value: Optional[float]) -> str:
    if value is None:
        return "N/A"
    return f"{value * 100:.2f}%"


def contains_match_norm(norm_series: pd.Series, keyword: str) -> pd.Series:
    norm_kw = normalize_text(keyword)
    if not norm_kw:
        return pd.Series([False] * len(norm_series), index=norm_series.index)
    return norm_series.fillna("").astype(str).str.contains(re.escape(norm_kw), regex=True)


def get_sender_display_name(ctx: Dict[str, Any]) -> str:
    first_name = (ctx.get("sender_first_name") or "").strip()
    last_name = (ctx.get("sender_last_name") or "").strip()
    username = (ctx.get("sender_username") or "").strip()

    telegram_name = f"{last_name}{first_name}".strip() or first_name
    if telegram_name:
        return telegram_name
    if username:
        return f"@{username}"
    return "사용자"


def extract_message_context(update: dict) -> Dict[str, Any]:
    msg = update.get("message") or update.get("edited_message")
    if not msg:
        return {
            "chat_id": None,
            "text": None,
            "sender_user_id": None,
            "sender_first_name": None,
            "sender_last_name": None,
            "sender_username": None,
            "document": None,
        }

    chat = msg.get("chat", {}) or {}
    sender = msg.get("from", {}) or {}
    document = msg.get("document")

    doc_info = None
    if document:
        doc_info = {
            "file_id": document.get("file_id"),
            "file_name": document.get("file_name"),
            "mime_type": document.get("mime_type"),
            "file_size": document.get("file_size"),
        }

    return {
        "chat_id": chat.get("id"),
        "text": msg.get("text") or msg.get("caption") or "",
        "sender_user_id": sender.get("id"),
        "sender_first_name": sender.get("first_name"),
        "sender_last_name": sender.get("last_name"),
        "sender_username": sender.get("username"),
        "document": doc_info,
    }


def notify_owner_of_external_query(owner_chat_id: str, ctx: Dict[str, Any]) -> None:
    try:
        sender_user_id = ctx.get("sender_user_id")
        text = (ctx.get("text") or "").strip()

        if not sender_user_id or not text:
            return
        if str(sender_user_id) == str(owner_chat_id):
            return
        if text == "/help":
            return
        if not (text.startswith("/조회") or text.startswith("/분석")):
            return

        display_name = get_sender_display_name(ctx)
        username = (ctx.get("sender_username") or "").strip()
        username_text = f"@{username}" if username else "없음"

        msg = (
            "[외부 사용자 질의 감지]\n"
            f"- 이름: {display_name}\n"
            f"- Username: {username_text}\n"
            f"- Telegram ID: {sender_user_id}\n"
            f"- 메시지: {text}"
        )
        #send_message(int(owner_chat_id), msg)
    except Exception:
        logging.exception("notify_owner_of_external_query failed")


def get_kst_today_str() -> str:
    return datetime.now(ZoneInfo("Asia/Seoul")).strftime("%Y-%m-%d")


def get_kst_today_year() -> int:
    return int(datetime.now(ZoneInfo("Asia/Seoul")).strftime("%Y"))


def load_question_limit_data() -> Dict[str, Any]:
    if not os.path.exists(config.QUESTION_LIMIT_FILE):
        return {}
    try:
        with open(config.QUESTION_LIMIT_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        logging.exception("question limit file load failed")
        return {}


def save_question_limit_data(data: Dict[str, Any]) -> None:
    tmp = config.QUESTION_LIMIT_FILE + ".tmp"
    with open(tmp, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    shutil.move(tmp, config.QUESTION_LIMIT_FILE)


def check_and_increment_question_limit(
    sender_user_id: Optional[int],
    limit: int
) -> Tuple[bool, int]:
    if sender_user_id is None:
        return True, 0

    if config.OWNER_CHAT_ID and str(sender_user_id) == str(config.OWNER_CHAT_ID):
        return True, 0

    today = get_kst_today_str()
    key = str(sender_user_id)

    usage_data = load_question_limit_data()

    if usage_data.get("_date") != today:
        usage_data = {"_date": today, "users": {}}

    users = usage_data.setdefault("users", {})
    used_count = int(users.get(key, 0))

    if used_count >= limit:
        return False, used_count

    users[key] = used_count + 1
    save_question_limit_data(usage_data)
    return True, used_count + 1


# =========================================================
# 상세 PDF
# =========================================================
def export_project_sheet_pdf(workbook_path: str, sheet_name: str, out_pdf: str) -> str:
    import pythoncom
    import win32com.client as win32

    pythoncom.CoInitialize()
    excel = None
    wb = None

    try:
        excel = win32.DispatchEx("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        excel.ScreenUpdating = False
        excel.EnableEvents = False

        wb = excel.Workbooks.Open(os.path.abspath(workbook_path), ReadOnly=True)
        ws = wb.Worksheets(sheet_name)
        ws.Activate()

        ws.ExportAsFixedFormat(
            Type=0,
            Filename=os.path.abspath(out_pdf),
            Quality=0,
            IncludeDocProperties=True,
            IgnorePrintAreas=False,
            OpenAfterPublish=False
        )
        return out_pdf

    finally:
        try:
            if wb is not None:
                wb.Close(SaveChanges=False)
        except Exception:
            pass
        try:
            if excel is not None:
                excel.Quit()
        except Exception:
            pass
        try:
            pythoncom.CoUninitialize()
        except Exception:
            pass


def _find_sheet_name(workbook_path: str, project_id: str) -> str:
    from openpyxl import load_workbook

    wb = load_workbook(workbook_path, read_only=True, keep_vba=True, data_only=True)
    try:
        if project_id in wb.sheetnames:
            return project_id

        norm_target = normalize_text(project_id)
        for s in wb.sheetnames:
            if normalize_text(s) == norm_target:
                return s

        raise ValueError(f"시트를 찾을 수 없습니다: {project_id}")
    finally:
        wb.close()


def export_project_pdf(project_id: str) -> str:
    if not os.path.exists(config.DETAIL_XLSX):
        raise FileNotFoundError(f"상세 엑셀 파일이 없습니다: {config.DETAIL_XLSX}")

    sheet_name = _find_sheet_name(config.DETAIL_XLSX, project_id)
    fd, out_pdf = tempfile.mkstemp(prefix=f"detail_{project_id}_", suffix=".pdf")
    os.close(fd)
    return export_project_sheet_pdf(config.DETAIL_XLSX, sheet_name, out_pdf)